# 找出两个数组中的公共连续子数组
def findCommonContinuousElements(logistics_route, commuting_route):
    result = []
    i = 0
    # 遍历A数组
    while i < len(logistics_route):
        j = 0
        # 遍历B数组
        while j < len(commuting_route):
            # 如果A[i] == B[j]，则开始查找公共连续子数组
            if logistics_route[i] == commuting_route[j]:
                start_i = i
                # 从A[i]和B[j]开始向后遍历，直到元素不相等
                while i < len(logistics_route) and j < len(commuting_route) and logistics_route[i] == commuting_route[j]:
                    i += 1
                    j += 1
                # 如果公共连续子数组长度大于1，则将其添加到结果中
                if i - start_i > 1:
                    result.append(logistics_route[start_i:i])
                # 这里要注意，因为上面的while循环中i和j都自增了1，所以这里要回退
                i -= 1 
                j -= 1 
            j += 1
        i += 1
    return result


# 找出commuting_route中在logistics_route中跨度最大的两个元素
def findMaxSpanElements(logistics_route, commuting_route):
    # 确定commuting_route中元素在logistics_route中的索引位置
    indexes = []
    for element in commuting_route:
        if element in logistics_route:
            indexes.append(logistics_route.index(element))
        else:
            indexes.append(None)

    # 排除无效索引（即logistics_route中不存在的元素）
    valid_indexes = [index for index in indexes if index is not None]

    # 找到索引范围的最小值和最大值
    if len(valid_indexes) == 0:
        return []
    else:
        start_index = min(valid_indexes)
        end_index = max(valid_indexes)

        # 提取logistics_route中对应索引范围的子集
        result = logistics_route[start_index:end_index+1]
        return result

# 将数组中的元素逆序，例如[1, 2, 3, 4] -> [4, 3, 2, 1]
def reverseArray(array):
    i = 0
    j = len(array) - 1
    while i < j:
        array[i], array[j] = array[j], array[i]
        i += 1
        j -= 1
    return array

# 找出两个数组中的公共连续子数组，考虑正反向
def findCommonContinuousElementsBidirection(logistics_route, commuting_route):
    result = []
    # 正向查找
    result.extend(findCommonContinuousElements(logistics_route, commuting_route))
    # 反向查找
    result.extend(findCommonContinuousElements(logistics_route, reverseArray(commuting_route)))
    return result   

# 找到每条通勤路径与每条物流路径的公共网格
# 输入：logistics_task_route.xlsx, commuting_route.xlsx
# 输出：commuting_logistics_common_route.xlsx
def findCommonRoute():
    # 打开logistics_task_route.xlsx和comuting_route.xlsx文件，读取数据
    import openpyxl
    workbook_logistics = openpyxl.load_workbook('dataSet/logistics_task_route.xlsx')
    logistics_sheet = workbook_logistics['logistics_route_grid_20']

    workbook_commuting = openpyxl.load_workbook('dataSet/commuting_route.xlsx')
    commuting_sheet = workbook_commuting['commuting_route_grid']

    workbook_common_route = openpyxl.load_workbook('dataSet/logistics_commuting_common_route.xlsx')
    common_sheet = workbook_common_route['common_route_grid_20']

    # 读取logistics_route表格中的第7列数据
    for i in range(2, logistics_sheet.max_row + 1):
        str_logistics_route = logistics_sheet.cell(row = i, column = 7).value
        logistics_route = eval(str_logistics_route)

        # 读取commuting_route表格中的第2列数据
        for j in range(2, commuting_sheet.max_row + 1):
            str_commuting_route = commuting_sheet.cell(row = j, column = 2).value
            commuting_route = eval(str_commuting_route)
            # 找出两个数组中的公共连续子数组
            common_route = findCommonContinuousElementsBidirection(logistics_route, commuting_route)
            # 将common_route转成一个字符串
            common_route_str = str(common_route)
            # 将common_route_str写入commuting_logistics_common_route.xlsx文件的common_route_grid表中的第j行第i列
            common_sheet.cell(row = j, column = i).value = common_route_str

    # 保存commuting_logistics_common_route.xlsx文件
    workbook_common_route.save('dataSet/logistics_commuting_common_route.xlsx')
    # 关闭文件
    workbook_logistics.close()
    workbook_commuting.close()
    workbook_common_route.close()
    
def findMaxspanRoute():
    # 打开logistics_task_route.xlsx和comuting_route.xlsx文件，读取数据
    import openpyxl
    workbook_logistics = openpyxl.load_workbook('dataSet/logistics_task_route.xlsx')
    logistics_sheet = workbook_logistics['logistics_route_grid_10']

    workbook_commuting = openpyxl.load_workbook('dataSet/commuting_route.xlsx')
    commuting_sheet = workbook_commuting['commuting_route_grid']

    workbook_maxspan_route = openpyxl.load_workbook('dataSet/logistics_commuting_maxspan_route.xlsx')
    maxspan_sheet = workbook_maxspan_route['maxspan_route_grid_10']

    # 读取logistics_route表格中的第7列数据
    for i in range(2, logistics_sheet.max_row + 1):
        str_logistics_route = logistics_sheet.cell(row = i, column = 7).value
        logistics_route = eval(str_logistics_route)

        # 读取commuting_route表格中的第2列数据
        for j in range(2, commuting_sheet.max_row + 1):
            str_commuting_route = commuting_sheet.cell(row = j, column = 2).value
            commuting_route = eval(str_commuting_route)
            # 找出两个数组中跨度最大的数组
            maxspan_route = findMaxSpanElements(logistics_route, commuting_route)
            # 将maxspan_route转成一个字符串
            maxspan_route_str = str(maxspan_route)
            # 将maxspan_route_str写入commuting_logistics_maxspan_route.xlsx文件的maxspan_route_grid表中的第j行第i列
            maxspan_sheet.cell(row = j, column = i).value = maxspan_route_str

    # 保存commuting_logistics_maxspan_route.xlsx文件
    workbook_maxspan_route.save('dataSet/logistics_commuting_maxspan_route.xlsx')
    # 关闭文件
    workbook_logistics.close()
    workbook_commuting.close()
    workbook_maxspan_route.close()




## 测试
if __name__ == '__main__':
    # logistics_route = [1079, 1080, 1081, 1026, 971, 916, 915, 860]
    # logistics_route = [946, 947, 948, 949, 950, 951, 952, 1007, 1008, 1009, 1010, 955, 956, 957, 958, 959, 960, 1015, 1016, 1017, 1018, 1019, 1020, 1021, 1076, 1077, 1078, 1079]
    # commuting_route1 = [976, 977, 976, 975, 974, 975, 1030, 1085, 1084, 1028, 1027, 1026, 1081, 1080, 1079, 1078, 1077, 1076, 1021, 1020, 1019, 1018, 1017, 1016, 1015, 960, 959, 958, 957, 956, 955, 1010, 1009, 1008, 1007, 952, 951, 950, 949, 948, 947, 1002, 1003]
    # commuting_route2 = [1814, 1759, 1814, 1869, 1814, 1759, 1704, 1649, 1648, 1647, 1646, 1591, 1536, 1481, 1426, 1370, 1315, 1260, 1259, 1204, 1203, 1202, 1201, 1200, 1199, 1143, 1142, 1087, 1086, 1085, 1084, 1028, 1027, 1026, 1081, 1080, 1079, 1078, 1077, 1076, 1021, 1020, 1019, 1018, 1017, 1016, 1015, 1071]
    # commuting_route3 = [1698, 1753, 1698, 1697, 1642, 1641, 1640, 1639, 1584, 1529, 1474, 1419, 1364, 1309, 1254, 1199, 1143, 1142, 1087, 1086, 1085, 1084, 1028, 1027, 1026, 1081, 1080, 1079, 1078, 1077, 1076, 1021, 1020, 1019, 1018, 1017, 962, 907, 852, 851, 906, 907]
    logistics_route = [1079, 1080, 1081, 1026, 971, 916, 915, 860]
    commuting_route = [1698, 1753, 1698, 1753, 1808, 1809, 1864, 1865, 1866, 1867, 1868, 1923, 1868, 1867, 1866, 1865, 1864, 1809, 1808, 1807, 1806, 1805, 1804, 1749, 1748, 1747, 1746, 1745, 1744, 1743, 1798, 1743, 1688, 1687, 1632, 1577, 1522, 1521, 1466, 1411, 1356, 1301, 1246, 1191, 1136, 1081, 1026, 971, 970, 915, 860, 805, 750, 749, 694, 639, 638, 583, 528, 527, 472, 417, 416, 361, 306, 251, 250, 251, 252, 307, 362, 307, 306, 305, 304, 303, 302, 356, 355, 354, 353, 352, 351, 350, 349, 348, 347, 346, 345, 344, 343, 398, 453, 452, 451]

    # print(findCommonContinuousElementsBidirection(logistics_route, commuting_route))
    # findCommonRoute()
    findMaxspanRoute()

    # logistics_route = [1, 2, 3, 4, 5, 6]
    # commuting_route = [1, 8, 5]

