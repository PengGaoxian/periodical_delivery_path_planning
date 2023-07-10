import heapq

# 定义节点类
class Node:
    def __init__(self, index, parent=None):
        self.index = index
        self.parent = parent
        self.g = 0  # 起始点到当前节点的实际代价
        self.h = 0  # 当前节点到目标节点的预估代价
        self.f = 0  # 综合代价（f = g + h）

    def __lt__(self, other):
        return self.f < other.f


# 定义 A* 算法函数
def astar_search(start, goal, graph):
    # 获取起始节点和目标节点
    start_node = Node(start)
    goal_node = Node(goal)

    # 初始化开放列表和关闭列表
    open_list = []
    closed_list = set()

    # 将起始节点添加到开放列表
    heapq.heappush(open_list, start_node)

    # 开始搜索
    while open_list:
        # 从开放列表中获取 f 值最小的节点
        current_node = heapq.heappop(open_list)

        # 将当前节点添加到关闭列表
        closed_list.add(current_node.index)

        # 判断是否到达目标节点
        if current_node.index == goal_node.index:
            path = []
            while current_node:
                path.append(current_node.index)
                current_node = current_node.parent
            return path[::-1]  # 返回反转的路径

        # 生成当前节点的相邻节点
        neighbors = generate_neighbors(current_node, graph)

        for neighbor in neighbors:
            # 如果相邻节点已在关闭列表中，则跳过
            if neighbor.index in closed_list:
                continue

            # 计算相邻节点的实际代价
            g = current_node.g + graph[current_node.index][neighbor.index]

            # 如果相邻节点不在开放列表中，则添加到开放列表
            if neighbor not in open_list:
                neighbor.g = g
                neighbor.h = calculate_heuristic(neighbor.index, goal_node.index)
                neighbor.f = neighbor.g + neighbor.h
                neighbor.parent = current_node
                heapq.heappush(open_list, neighbor)
            else:
                # 如果相邻节点已在开放列表中且新的实际代价更小，则更新代价和父节点
                if g < neighbor.g:
                    neighbor.g = g
                    neighbor.f = neighbor.g + neighbor.h
                    neighbor.parent = current_node

    # 如果没有找到路径，返回空列表
    return []


# 定义生成相邻节点的函数
def generate_neighbors(node, graph):
    neighbors = []
    for index in graph[node.index]:
        neighbor = Node(index)
        neighbors.append(neighbor)
    return neighbors


# 定义计算启发式代价的函数（这里使用曼哈顿距离作为启发式函数）
def calculate_heuristic(index, goal):
    x_index = index % 55
    y_index = index // 55
    x_goal = goal % 55
    y_goal = goal // 55
    return abs(x_index - x_goal) + abs(y_index - y_goal)


# # 测试示例
# graph = {
#     0: {1: 1, 2: 1},
#     1: {0: 1, 3: 1},
#     2: {0: 1, 3: 1},
#     3: {1: 1, 2: 1, 4: 1},
#     4: {3: 1, 5: 1},
#     5: {4: 1}
# }
# 
# start = 0
# goal = 5
# 
# path = astar_search(start, goal, graph)
# print("最短路径：", path)


# # 打开area_grid_Neighbors.dbf文件，读取数据
# from dbfread import DBF
# 
# # 指定要读取的 DBF 文件路径
# dbf_file = 'area_grid_Neighbors.dbf'
# 
# # 使用 DBF 函数打开 DBF 文件
# table = DBF(dbf_file)
# 
# # 生成graph
# graph = {}
# 
# for record in table:
#     src_uid = record['src_uid']
#     neighbor = record['nbr_uid']
# 
#     graph.setdefault(src_uid, {})[neighbor] = 1
# 
# print(graph)
# 
# start = 946
# goal = 1023
# 
# path = astar_search(start, goal, graph)
# print("最短路径：", path)
if __name__ == '__main__':
    # 打开area_grid_Neighbors.xlsx文件，读取数据
    import openpyxl
    workbook = openpyxl.load_workbook('dataSet/area_grid_Neighbors.xlsx')
    sheet = workbook['area_grid_Neighbors_route']
    # 遍历sheet中的行数据，从第二行开始，并将其存储在graph中
    graph = {}
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3):
        # 将row中的第二列数据存入src_uid, 将第三列数据存入neighbor
        src_uid = row[1].value
        neighbor = row[2].value
        # 将src_uid作为key, neighbor作为value存入graph中
        graph.setdefault(src_uid, {})[neighbor] = 1

    # print(graph)

    start = 860
    goal = 946

    path = astar_search(start, goal, graph)
    print("最短路径：", path)
