# 从xlsx文件中随机算出指定个数的行数据，写入xlsx文件的表中
# 输入：xlsx文件名xlsx_name，行数row_num
# 输出：xlsx文件中的某一个表
import openpyxl
import random

def select_rows_random(xlsx_name, row_num):
    # 打开xlsx文件
    workbook = openpyxl.load_workbook(xlsx_name)
    # 获取xlsx文件中的sheet
    sheet_inborder = workbook['WGS84_in_border']
    sheet_inborder_filter = workbook['WGS84_in_border_filter'] 
    # 获取sheet中除表头外的最大行数
    max_row = sheet_inborder.max_row
    # 生成row_num个随机数
    random_list = random.sample(range(2, max_row + 1), row_num)

    # 读取sheet中对应的random_list所在的行，写入xlsx文件的WGS84_inborder表中
    for index, line in enumerate(random_list):
        source_row = sheet_inborder[line]

        for cell in source_row:
            column_num = cell.column
            target_cell = sheet_inborder_filter.cell(row=index + 2, column=column_num)
            target_cell.value = cell.value

    # 保存工作簿到xlsx文件中
    workbook.save(xlsx_name)
    # 关闭工作簿
    workbook.close()

if __name__ == '__main__':
    # 从xlsx文件中随机算出指定个数的行数据，写入xlsx文件的表中
    select_rows_random('inborder_ArcGIS_home.xlsx', 60)
    select_rows_random('inborder_ArcGIS_company.xlsx', 25)
    select_rows_random('inborder_ArcGIS_warehouse.xlsx', 12)