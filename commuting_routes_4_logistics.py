# 判断两个一维数组是否相等
def is_array_equal(array1, array2):
    if len(array1) != len(array2):
        return False
    for i in range(len(array1)):
        if array1[i] != array2[i]:
            return False
    return True


# 判断二维数组lit_2d_array是否包含于二维数组big_2d_array中
def is_2d_array_contained_in_2d_array(lit_2d_array, big_2d_array):
    # 遍历array_2d_B中的每一个元素b
    for l in lit_2d_array:
        equal_flag = False
        for b in big_2d_array:
            equal_flag = is_array_equal(l, b)
            if equal_flag:
                break

        if not equal_flag:
           break

    return equal_flag 
# 测试用例
# lit = [[1, 2, 3], [4, 5, 6]]
# big = [[1, 2, 3], [4, 5, 5], [7, 8, 9]]
# print(is_2d_array_contained_in_2d_array(lit, big))


# 判断一个二维数组big_2d_array是否包含一个三维数组lit_3d_array中的任意一个数组
def is_any_3d_array_contained_in_2d_array(lit_3d_array, big_2d_array):
    # 遍历lit_3d_array中的每一个二维数组
    for lit_2d_array in lit_3d_array:
        # 如果lit_2d_array包含于big_2d_array中，则返回True
        if is_2d_array_contained_in_2d_array(lit_2d_array, big_2d_array):
            return True
    return False
# 测试用例
# big_2d_array = [[1, 2, 3], [4, 5, 6], [7, 8, 9]]
# lit_2d_array1 = [[1, 2, 3], [4, 5, 6]]
# lit_2d_array2 = [[3, 4, 6],[7, 8, 9]]
# lit_2d_array3 = [[2, 3, 4, 7], [5, 6, 7, 8]]
# lit_3d_array = [lit_2d_array1, lit_2d_array2, lit_2d_array3]
# print(is_any_3d_array_contained_in_2d_array(lit_3d_array, big_2d_array))


# 对公共路径组合进行排序，按照物流路径的先后顺序
def sort_combination(combination_2d, logistics_route_1d_array):
    first_index_list = []
    last_index_list = []
    # 遍历combination_2d中的每一个数组，为common_1d_array中第一个元素和最后一个元素在logistics_route_1d_array中的索引分别添加到first_index_list和last_index_list中
    for common_1d_array in combination_2d:
        # 将common_1d_array中第一个元素在logistics_route_1d_array中的索引添加到first_index_list中
        first_index_list.append(logistics_route_1d_array.index(common_1d_array[1]))
        # 将common_1d_array中最后一个元素在logistics_route_1d_array中的索引添加到last_index_list中
        last_index_list.append(logistics_route_1d_array.index(common_1d_array[-1]))
    
    # 对first_index_list进行升序排列
    first_index_list_sorted = sorted(first_index_list)
    # 对last_index_list，按照first_index_list_sorted的调整顺序进行调整
    last_index_list_sorted = [last_index_list[first_index_list.index(first_index)] for first_index in first_index_list_sorted]
    # 从first_index_list_sorted的第3个元素开始，如果第i个元素大于等于last_index_list_sorted中的第i-1个元素，则输出combination_d_sorted，否则输出[]
    for i in range(1, len(first_index_list_sorted)-1):
        if first_index_list_sorted[i+1] > last_index_list_sorted[i]:
            return []

    # 对combination_2d，按照first_index_list_sorted的调整顺序进行调整
    combination_2d_sorted = [combination_2d[first_index_list.index(first_index)] for first_index in first_index_list_sorted]

    return combination_2d_sorted
# 测试用例
# logistics_route_1d_array= [1079, 1080, 1081, 1026, 971, 916, 915, 860]
# combination_2d = ([9937, 1026, 971, 916, 915], [9942, 1079, 1080, 1081, 1026], [9987, 915, 860])
# print(sort_combination(combination_2d, logistics_route_1d_array))


import itertools
# 从commuting_route中选择数组的组合，其中任一组合都能完成logistics_route
def get_combinations(logistics_route_1d_array, commuting_route_2d_array, max_combination_num):
    max_combination_capacity = 20
    result_3d_array = [] # [[[b1, b2], [b1, b3]],[[b1, b3], [b2, b3]]]
    # 从commuting_route中选择r个数组的组合
    for r in range(1, max_combination_num + 1): # r = [[1,2,3],[4,5,6]]
        # 获取r个数组的组合
        combinations = list(itertools.combinations(commuting_route_2d_array, r)) # [(b1, b2), (b1, b3), (b2, b3)]
        # 遍历combinations中的每一个组合combo
        for combo in combinations:
            # 如果combo包含result_3d_array中的任意一个数组，则跳过combo
            ignore_flag = is_any_3d_array_contained_in_2d_array(result_3d_array, combo)
            if ignore_flag:
                continue

            # 将数组combo中的数组合并成一个数组
            combined = [num for sublist in combo for num in sublist]
            # 如果数组combined包含数组logistics_route_1d_array中的所有元素，则将数组combo添加到result中
            if set(logistics_route_1d_array).issubset(set(combined)):
                # 对数组combo按照logistics_route_1d_array的先后顺序进行排序
                combination_sorted = sort_combination(combo, logistics_route_1d_array)
                # 如果combination_sorted不为空，且result_3d_array中的元素不超过max_combination_num，则将combination_sorted添加到result_3d_array中
                if combination_sorted != [] and len(result_3d_array) < max_combination_capacity:
                    result_3d_array.append(combination_sorted)
                elif len(result_3d_array) >= max_combination_capacity:
                    break
    return result_3d_array
# 测试用例
# logistics_route_1d_array= [1079, 1080, 1081, 1026, 971, 916, 915, 860]
# commuting_route_2d_array = [['task11', 860], ['task12', 860], ['task13', 860], ['task14', 1079, 1080, 1081, 1026], ['task37', 1026, 971, 916, 915], ['task39', 860], ['task42', 1079, 1080, 1081, 1026], ['task43', 860], ['task53', 1079, 1080, 1081, 1026], ['task54', 860], ['task55', 1081, 1026, 971, 916, 915, 860], ['task57', 860], ['task60', 860], ['task65', 1081, 1026, 971, 916, 915, 860], ['task68', 1081, 1026, 971, 916, 915, 860], ['task87', 915, 860], ['task89', 860]]
# print(get_combinations(logistics_route_1d_array, commuting_route_2d_array))


# 获取物流路径的候选通勤路径的候选集合
def generate_candidate_route_set():
    # 打开logistics_commuting_route.xlsx文件的maxspan_route_grid表
    import openpyxl
    wb_logistics_commuting_route = openpyxl.load_workbook('dataSet/logistics_commuting_maxspan_route.xlsx')
    commuting_sheet = wb_logistics_commuting_route['maxspan_route_grid_10']
    # 打开logistics_task_route.xlsx文件的logistics_route_grid表
    wb_logistics_task_route = openpyxl.load_workbook('dataSet/logistics_task_route.xlsx')
    logistics_sheet = wb_logistics_task_route['logistics_route_grid_10']
    # 打开candidate_commuting_route_set.xlsx文件的logistics_task表
    wb_candidate_commuting_route_set = openpyxl.load_workbook('dataSet/candidate_commuting_route_set.xlsx')
    candidate_sheet = wb_candidate_commuting_route_set['logistics_task_10']

    # 从第2行开始，遍历logistics_sheet中的每一行
    for logistics_row in range(2, logistics_sheet.max_row + 1):
        # 将row行的第7列数据存放到logistics_route_1d_array中
        logistics_route_1d_array_str = logistics_sheet.cell(row=logistics_row, column=7).value
        # 将logistics_route_1d_array由字符串类型转为数组类型
        logistics_route_1d_array = eval(logistics_route_1d_array_str)

        # 设置最小的通勤路径网格数为物流路径网格数的1/5，如果最小的通勤路径网格数小于2，则设置最小的通勤路径网格数为2
        min_route_grid_num = 2

        commuting_route_2d_array = []
        combination_3d_array = []
        # 遍历commuting_sheet中的每一行
        for commuting_row in range(2, commuting_sheet.max_row + 1):
            # 获取commuting_row行中的第row列数据
            commuting_route_1d_array_str = commuting_sheet.cell(row=commuting_row, column=logistics_row).value
            # 将commuting_route_1d_array由字符串类型转为数组类型
            commuting_route_1d_array = eval(commuting_route_1d_array_str)
            # 如果commuting_route_1d_array的元素个数大于1个，则将commuting_route_1d_array添加到commuting_route_2d_array中
            if commuting_route_1d_array.__len__() >= min_route_grid_num:
                commuting_route_1d_array.insert(0, 'C' + str(commuting_row - 1))
                commuting_route_2d_array.append(commuting_route_1d_array)

        # 设置通勤路径组合数量的上限（设置大了运行慢）
        max_combination_num = 5
        # 计算能完成logistics_route_1d_array的commuting_route_2d_array的组合
        combination_3d_array = get_combinations(logistics_route_1d_array, commuting_route_2d_array, max_combination_num)
        # 设置写入candidate_sheet表的行号
        write_candidate_row = 2
        # 将combination_3d_array写入到candidate_sheet的第logistics_row-1列
        for combination_2d_array in combination_3d_array:
            # 将combination_2d_array转为字符串类型
            combination_2d_array_str = str(combination_2d_array)
            print(str(logistics_row - 1) + ' ' + str(write_candidate_row - 1) + ' ' + str(combination_2d_array))

            # 将combination_2d_array_str以追加的形式写入到candidate_sheet的第logistics_row-1列，从第2行开始依次往下写
            candidate_sheet.cell(row=write_candidate_row, column=logistics_row - 1).value = combination_2d_array_str
            write_candidate_row += 1

    # 保存candidate_commuting_route_set.xlsx文件
    wb_candidate_commuting_route_set.save('dataSet/candidate_commuting_route_set.xlsx')
    # 关闭logistics_commuting_route.xlsx文件
    wb_logistics_commuting_route.close()
    # 关闭logistics_task_route.xlsx文件
    wb_logistics_task_route.close()
    # 关闭candidate_commuting_route_set.xlsx文件
    wb_candidate_commuting_route_set.close()

if __name__ == '__main__':
    generate_candidate_route_set()