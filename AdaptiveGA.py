import openpyxl
# 车辆类
class VehSeqComp:
    # 初始化
    def __init__(self, comp_3d_array):
        # 个体基因
        self.comp_3d_array = comp_3d_array
        self.gene = []
        self.singleTaskCost_1d_array = []
        self.compCost = 0
        self.IsinTimeLimit = 0
        self.IsinWeightLimit = 0

# 编码
def encode(filePath, tableName, comp_3d_array):
    # 打开excel文件
    wb = openpyxl.load_workbook(filePath)
    # 获取sheet
    sheet = wb[tableName]
    # 存储编码后的变量
    encode_list = []
    # 遍历comp_3d_array
    for index, vehicleSeq4SingleTask_2d_array in enumerate(comp_3d_array):
        # 遍历sheet中的第一列数据
        for cell in sheet.iter_rows(min_row=2, min_col=index+1, max_col=index+1):
            value_str = cell[0].value
            vehSeqST_2d_array_str = str(vehicleSeq4SingleTask_2d_array)
            if value_str == vehSeqST_2d_array_str:
                encode_list.append(cell[0].row-1)
                break
    return encode_list
            

# 解码
def decode(filePath, tableName, encode):
    # 打开excel文件
    wb = openpyxl.load_workbook(filePath)
    # 获取sheet
    sheet = wb[tableName]
    # 存储解码后的变量
    decode_list = []
    # 遍历encode_3d_array
    for index, encode_value in enumerate(encode):
        # 遍历sheet中的第index列第encode_value行单元格的数据
        for cell in sheet.iter_rows(min_row=encode_value+1, min_col=index+1, max_col=index+1):
            value_str = cell[0].value
            decode_list.append(eval(value_str))
            break
    return decode_list


# 初始化种群
import random
from extract_parameter import getEncodeNumBoundary, getVehicleCommutingTime, getLogisticsTaskInfo, getRelayPoint
def initPopulation(encodeNumBoundary_list):
    # 种群
    population = []
    # 种群大小
    populationSize = 50
    # 随机生成种群
    for i in range(populationSize):
        individual_gene = []
        # 获取物流任务的候选车辆序列集大小
        for j in range(len(encodeNumBoundary_list)):
            # 随机生成最小值为0，最大值为encodeNumBoundary_list[j]-1的整数
            individual_gene.append(random.randint(1, encodeNumBoundary_list[j]))

        # 将个体添加到种群中
        population.append(individual_gene)

    # 返回种群
    return population

# 计算种群适应度
def calPopulationFitness(population, commuting_time_1d_array, logisticsTaskInfo_list, candidate_filePath, candidate_tableName):
    population_result = []
    # 遍历种群
    for individual_gene in population:
        # 解码获得车辆序列组合
        comp_3d_array = decode(candidate_filePath, candidate_tableName, individual_gene)
        # 获取车辆序列组合的接力点
        relay_points_2d_array = getRelayPoint('dataSet/trans_points.xlsx', '100%', comp_3d_array)
        # 计算comp_3d_array的成本
        comp_result = calIndividualFitness(comp_3d_array, commuting_time_1d_array, logisticsTaskInfo_list, relay_points_2d_array, candidate_filePath, candidate_tableName)
        # 将comp_result存储到population_result中
        population_result.append(comp_result)
    return population_result


# 计算个体适应度
import numpy as np
from costLimit import isMeetTimeLimit4VehicleSeqComp, isMeetWeightLimit4VehicleSeqComp, calDetourCost4SingleTask, calTransportCost4SingleTask 
def calIndividualFitness(comp_3d_array, commuting_time_1d_array, logisticsTaskInfo_list, relay_points_2d_array, candidate_filePath, candidate_tableName):
    # 将个体基因解码成车辆序列组合
    np_logisticsTaskInfo_list = np.array(logisticsTaskInfo_list)
    # 物流任务的起止点为logisticsTaskInfo_list的第二列和第三列
    logistics_start_end_2d_array = np_logisticsTaskInfo_list[:, 1:3]
    # 物流任务的重量为logisticsTaskInfo_list的第一列
    logistics_task_weight_1d_array = np_logisticsTaskInfo_list[:, 0]

    # 个体适应度
    comp_result = VehSeqComp(comp_3d_array)
    comp_result.gene = encode(candidate_filePath, candidate_tableName, comp_3d_array)

    # 计算车辆序列组合是否满足时间限制和载重限制
    time_flag = isMeetTimeLimit4VehicleSeqComp(comp_3d_array, commuting_time_1d_array, logistics_start_end_2d_array, relay_points_2d_array)
    weight_flag = isMeetWeightLimit4VehicleSeqComp(comp_3d_array, logistics_task_weight_1d_array)
    comp_result.IsinTimeLimit = time_flag
    comp_result.IsinWeightLimit = weight_flag

    if time_flag and weight_flag:
        # 遍历comp_3d_array
        for index, vehicleSeq4SingleTask_2d_array in enumerate(comp_3d_array):
            logistics_start_grid = logistics_start_end_2d_array[index][0]
            logistics_end_grid = logistics_start_end_2d_array[index][1]
            relay_point_1d_array = relay_points_2d_array[index]

            # 计算绕行成本
            detour_cost = calDetourCost4SingleTask(logistics_start_grid, logistics_end_grid, relay_point_1d_array, vehicleSeq4SingleTask_2d_array)

            # 计算运输成本
            transport_cost = calTransportCost4SingleTask(logistics_start_grid, logistics_end_grid, relay_point_1d_array, vehicleSeq4SingleTask_2d_array)

            # 将绕行成本和运输成本追加到comp_result.singleTaskCost_2d_array中
            trans_detour_cost = detour_cost + transport_cost
            comp_result.singleTaskCost_1d_array.append(trans_detour_cost)

            comp_result.compCost+= trans_detour_cost
    else:
        comp_result.compCost= 100000
    # 返回个体适应度
    return comp_result

# 选择操作，锦标赛选择
def selection(population_result, tournament_size):
    compCost = [individual_result.compCost for individual_result in population_result]
    # 从种群中选择一组父代个体
    selected_population = []
    # 种群大小
    population_size = len(compCost)

    while len(selected_population) < len(compCost):  # 假设要选择2个父代个体
        tournament_candidates = random.sample(range(population_size), tournament_size)  # 随机选择一组候选个体

        # 以tournament_candidate为索引，选出compCost适应度最高的个体
        winner = min(tournament_candidates, key=lambda x: compCost[x])

        selected_population.append(population_result[winner])

    return selected_population


# 交叉操作
def crossover(population):
    # 交叉后的种群
    newPopulation = []
    # 交叉概率
    crossoverProbability = 0.5
    # 交叉
    for i in range(len(population)):
        # 随机数
        r = random.random()
        # 交叉
        if r < crossoverProbability:
            # 交叉个体
            individual1 = population[i]
            # 交叉个体
            individual2 = random.choice(population)
            # 交叉点
            crossoverPoint = random.randint(0, len(individual1) - 1)
            # 交叉
            individual_new = individual1[0:crossoverPoint] + individual2[crossoverPoint:len(individual2)]
            # 添加到种群中
            newPopulation.append(individual_new)
        else:
            # 添加到种群中
            newPopulation.append(population[i])
    # 返回交叉后的种群
    return newPopulation

# 变异操作
def mutation(population, encodeNumBoundary_list):
    # 变异后的种群
    newPopulation = []
    # 变异概率
    mutationProbability = 0.2
    # 变异
    for i in range(len(population)):
        # 随机数
        r = random.random()
        # 变异
        if r < mutationProbability:
            # 变异个体
            individual = population[i]
            # 变异点
            mutationPoint = random.randint(0, len(individual) - 1)
            # 变异
            individual[mutationPoint] = random.randint(1, encodeNumBoundary_list[mutationPoint])
            # 添加到种群中
            newPopulation.append(individual)
        else:
            # 添加到种群中
            newPopulation.append(population[i])
    # 返回变异后的种群
    return newPopulation

# 功能：使用遗传算法计算能完成物流任务集中所有任务的车辆序列组合
# 输入：物流任务集，每一个物流任务的车辆序列集
# 输出：车辆序列组合及其对应的总成本
def AdaptiveGA(encodeNumBoundary_list, commuting_time_1d_array, logisticsTaskInfo_list, candidate_filePath, candidate_tableName):
    # 初始化种群
    population = initPopulation(encodeNumBoundary_list)
    # 计算种群适应度
    population_result = calPopulationFitness(population, commuting_time_1d_array, logisticsTaskInfo_list, candidate_filePath, candidate_tableName)
    # 迭代次数
    iterNum = 1
    # 最大迭代次数
    maxIterNum = 100
    # 最小车辆序列组合成本
    min_compCost = 100000
    # 最优个体和最优个体列表
    bestIndividualResult = None
    bestIndividualResult_list = []
    # 迭代
    while iterNum <= maxIterNum:
        # 输出迭代次数
        print("iterNum:", iterNum)
        # 选择
        selection_result = selection(population_result, 3)
        selection_population = [individual_result.gene for individual_result in selection_result]
        # 交叉
        crossover_population = crossover(selection_population)
        # 变异
        mutation_population = mutation(crossover_population, encodeNumBoundary_list)
        # 计算种群适应度
        population_result = calPopulationFitness(mutation_population, commuting_time_1d_array, logisticsTaskInfo_list, candidate_filePath, candidate_tableName)
        # 选择最优个体
        for individual_result in population_result:
            if individual_result.compCost < min_compCost:
                min_compCost = individual_result.compCost
                bestIndividualResult = individual_result

        bestIndividualResult_list.append(bestIndividualResult)
        # 迭代次数加一
        iterNum += 1
    # 返回最优个体及其适应度值列表
    return bestIndividualResult_list


if __name__ == '__main__':
    candidate_filePath = 'dataSet/candidate_commuting_route_set.xlsx'
    candidate_tableName = 'logistics_task_20'
    result_filePath = 'dataSet/comp_result_20.xlsx'

    encodeNumBoundary_list = getEncodeNumBoundary(candidate_filePath, candidate_tableName)
    commuting_time_1d_array = getVehicleCommutingTime('dataSet/commuting_route.xlsx', 'commuting_route_grid')
    logisticsTaskInfo_list = getLogisticsTaskInfo('dataSet/logistics_task.xlsx', 'WGS84_logistics_task_20')
    # 打开comp_result.xlsx文件的result表
    workbook = openpyxl.load_workbook(result_filePath)
    for i in range(5):
        bestIndividualResult_list = AdaptiveGA(encodeNumBoundary_list, commuting_time_1d_array, logisticsTaskInfo_list, candidate_filePath, candidate_tableName)
        worksheet = workbook['result' + str(i+1)]
        # 在worksheet中写入bestIndividualResult_list中每个个体的信息
        for j in range(len(bestIndividualResult_list)):
            worksheet.cell(row=j+2, column=1).value = str(bestIndividualResult_list[j].comp_3d_array)
            worksheet.cell(row=j+2, column=2).value = str(bestIndividualResult_list[j].gene)
            worksheet.cell(row=j+2, column=3).value = str(bestIndividualResult_list[j].singleTaskCost_1d_array)
            worksheet.cell(row=j+2, column=4).value = bestIndividualResult_list[j].compCost
            worksheet.cell(row=j+2, column=5).value = bestIndividualResult_list[j].IsinTimeLimit
            worksheet.cell(row=j+2, column=6).value = bestIndividualResult_list[j].IsinWeightLimit

    # 保存workbook
    workbook.save(result_filePath)
    # 关闭workbook
    workbook.close()
