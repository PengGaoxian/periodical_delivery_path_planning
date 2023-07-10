# 路径规划函数
# 输入参数：起点坐标、终点坐标、地图
# 输出参数：路径散点坐标

import json
from urllib import request
def getCommuteRoute():
    url = 'https://restapi.amap.com/v5/direction/driving?key=126486759bec7d1bd9fbd1d5e851744f&origin=112.945858,28.215530&destination=112.988784,28.191318&show_fields=cost,polyline'

    html = request.urlopen(url).read()
    js = json.loads(html)

    distance = js['route']['paths'][0]['distance']
    duration = js['route']['paths'][0]['cost']['duration']
    steps = js['route']['paths'][0]['steps']

    pts = []

    for p in steps:
        pts.extend(p['polyline'].split(';'))

    return pts, distance, duration

def getCommuteRoutes(xlsx_name, source_sheet, target_sheet):
    import openpyxl
    url = 'https://restapi.amap.com/v5/direction/driving?key=126486759bec7d1bd9fbd1d5e851744f&origin={0}&destination={1}&show_fields=cost,polyline'

    # 读取company_home.xlsx文件中的origin_destination表单中的origin列和destination列
    workbook = openpyxl.load_workbook(xlsx_name)
    sheet1 = workbook[source_sheet]
    sheet2 = workbook[target_sheet]
    max_row = sheet1.max_row

    pts = []
    ID = 1

    for i in range(2, max_row + 1):
        origin = sheet1.cell(row=i, column=9).value
        destination = sheet1.cell(row=i, column=10).value

        html = request.urlopen(url.format(origin, destination)).read()
        js = json.loads(html)

        distance = js['route']['paths'][0]['distance']
        duration = js['route']['paths'][0]['cost']['duration']
        steps = js['route']['paths'][0]['steps']

        for path in steps:
            for path_point in path['polyline'].split(';'):
                # 以追加的形式，按行将ID、path_point的lng、path_point的lat、distance、duration分别写入xlsx文件中的routes表单的第1、2、3、4列，每次写入一行
                lng = path_point.split(',')[0]
                lat = path_point.split(',')[1]
                sheet2.append([lng, lat, ID, distance, duration])
                
                pts.append([lng, lat, ID, distance, duration])

        print(ID)
        ID += 1

        workbook.save(xlsx_name)

    workbook.close()

    return pts

# area_grid_seqtable.xlsx文件的area_grid_seq_deduplicate表达中有两列分别是route_ID和route_grid_uid，将route_ID相同的route_grid_uid合并到一个数组中，然后将这个数组写入commuting_route.xlsx文件中的WGS84_commuting_route表单中
def getCommuteRoutesGrid(source_xlsx, target_xlsx, source_sheet, target_sheet):
    import openpyxl
    workbook_source = openpyxl.load_workbook(source_xlsx)
    workbook_target = openpyxl.load_workbook(target_xlsx) 
    sheetS = workbook_source[source_sheet]
    sheetT = workbook_target[target_sheet]
    max_row = sheetS.max_row

    grid_uid = []
    ID = 1

    for i in range(2, max_row + 1):
        route_ID = sheetS.cell(row=i, column=1).value
        route_grid_uid = sheetS.cell(row=i, column=2).value

        if route_ID == ID:
            grid_uid.append(route_grid_uid)
            if i == max_row:
                print(ID, grid_uid)
                # 将grid_uid数组转化为字符串
                str_grid_uid = str(grid_uid)
                # 将ID和str_grid_uid分别写入commuting_route.xlsx文件中的WGS84_commuting_route表单的第1、2列，每次写入一行
                sheetT.append([ID, str_grid_uid])
        else:
            print(ID, grid_uid)
            # 将grid_uid数组转化为字符串
            str_grid_uid = str(grid_uid)
            # 将ID和str_grid_uid分别写入commuting_route.xlsx文件中的WGS84_commuting_route表单的第1、2列，每次写入一行
            sheetT.append([ID, str_grid_uid])

            grid_uid = []
            grid_uid.append(route_grid_uid)
            ID += 1

    workbook_target.save(target_xlsx)
    workbook_source.close()
    workbook_target.close()

    return grid_uid 

if __name__ == '__main__':
    # pts, distance, duration = getCommuteRoute()
    # print(pts)
    # print(distance)
    # print(duration)
    
    # getCommuteRoutes('company_home_test.xlsx', 'GCJ02_origin_destination', 'GCJ02_route_points')

    getCommuteRoutesGrid('dataSet/area_grid_seqtable.xlsx', 'dataSet/commuting_route.xlsx', 'area_grid_seq_deduplicate', 'commuting_route_grid')