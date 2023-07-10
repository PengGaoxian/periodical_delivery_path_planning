# 将GCJ02坐标系的点转换为WGS84坐标系的点
# 输入：GCJ02坐标点的xlsx文件
# 输出：WGS84坐标点的xlsx文件

from coord_convert.transform import gcj2wgs, wgs2gcj, bd2gcj, gcj2bd
import openpyxl

def transfer(xlsx_name, transfer_type, source_sheet, target_sheet):
    # 打开xlsx文件
    workbook = openpyxl.load_workbook(xlsx_name)

    # 获取xlsx文件中的sheet
    source_sheet= workbook[source_sheet]
    target_sheet = workbook[target_sheet]

    if transfer_type == 'gcj02_to_wgs84':
        # 从gcj02_sheet中的第二行开始读取数据，转换后写入wgs84_sheet中，从第二行开始写入
        for i in range(2, source_sheet.max_row + 1):
            # 读取第i行的第1列到第2列
            gcj02_lng = source_sheet.cell(row=i, column=1).value
            gcj02_lat = source_sheet.cell(row=i, column=2).value
            # 将GCJ02坐标点转换为WGS84坐标点
            wgs84_lng, wgs84_lat = gcj02_to_wgs84(gcj02_lng, gcj02_lat)
            # 将WGS84坐标点写入xlsx文件中的WGS84表，从第2行开始写入
            target_sheet.cell(row=i, column=1).value = wgs84_lng
            target_sheet.cell(row=i, column=2).value = wgs84_lat

        # 保存工作簿到xlsx文件中
        workbook.save(xlsx_name)

        # 关闭工作簿
        workbook.close()
    elif transfer_type == 'wgs84_to_gcj02':
        # 从wgs84_sheet中的第二行开始读取数据，转换后写入gcj02_sheet中，从第二行开始写入
        for i in range(2, source_sheet.max_row + 1):
            # 读取第i行的第1列到第2列
            wgs84_lng = source_sheet.cell(row=i, column=1).value
            wgs84_lat = source_sheet.cell(row=i, column=2).value
            # 将WGS84坐标点转换为GCJ02坐标点
            gcj02_lng, gcj02_lat = wgs84_to_gcj02(wgs84_lng, wgs84_lat)
            # 将WGS84坐标点写入xlsx文件中的WGS84表，从第2行开始写入
            target_sheet.cell(row=i, column=1).value = gcj02_lng
            target_sheet.cell(row=i, column=2).value = gcj02_lat

        # 保存工作簿到xlsx文件中
        workbook.save(xlsx_name)

        # 关闭工作簿
        workbook.close()
    elif transfer_type == 'bd09_to_gcj02':
        # 从bd09_sheet中的第二行开始读取数据，转换后写入gcj02_sheet中，从第二行开始写入
        for i in range(2, source_sheet.max_row + 1):
            # 读取第i行的第1列到第2列
            bd09_lng = source_sheet.cell(row=i, column=1).value
            bd09_lat = source_sheet.cell(row=i, column=2).value
            # 将BD09坐标点转换为GCJ02坐标点
            gcj02_lng, gcj02_lat = bd09_to_gcj02(bd09_lng, bd09_lat)
            # 将WGS84坐标点写入xlsx文件中的WGS84表，从第2行开始写入
            target_sheet.cell(row=i, column=1).value = gcj02_lng
            target_sheet.cell(row=i, column=2).value = gcj02_lat

        # 保存工作簿到xlsx文件中
        workbook.save(xlsx_name)

        # 关闭工作簿
        workbook.close()
    elif transfer_type == 'gcj02_to_bd09':
        # 从gcj02_sheet中的第二行开始读取数据，转换后写入bd09_sheet中，从第二行开始写入
        for i in range(2, source_sheet.max_row + 1):
            # 读取第i行的第1列到第2列
            gcj02_lng = source_sheet.cell(row=i, column=1).value
            gcj02_lat = source_sheet.cell(row=i, column=2).value
            # 将GCJ02坐标点转换为BD09坐标点
            bd09_lng, bd09_lat = gcj02_to_bd09(gcj02_lng, gcj02_lat)
            # 将WGS84坐标点写入xlsx文件中的WGS84表，从第2行开始写入
            target_sheet.cell(row=i, column=1).value = bd09_lng
            target_sheet.cell(row=i, column=2).value = bd09_lat

        # 保存工作簿到xlsx文件中
        workbook.save(xlsx_name)

        # 关闭工作簿
        workbook.close()
        
def gcj02_to_wgs84(gcj02_lng, gcj02_lat):
    wgs84_lng, wgs84_lat = gcj2wgs(gcj02_lng, gcj02_lat)
    return wgs84_lng, wgs84_lat

def wgs84_to_gcj02(wgs84_lng, wgs84_lat):
    gcj02_lng, gcj02_lat = wgs2gcj(wgs84_lng, wgs84_lat)
    return gcj02_lng, gcj02_lat

def gcj02_to_bd09(gcj02_lng, gcj02_lat):
    bd09_lng, bd09_lat = gcj2bd(gcj02_lng, gcj02_lat)
    return bd09_lng, bd09_lat

def bd09_to_gcj02(bd09_lng, bd09_lat):
    gcj02_lng, gcj02_lat = bd2gcj(bd09_lng, bd09_lat)
    return gcj02_lng, gcj02_lat

if __name__ == '__main__':
    # # gcj02_to_wgs84函数测试
    # gcj_lng = 112.944247
    # gcj_lat = 27.853246
    # wgs84_lng, wgs84_lat = bd09_to_gcj02(gcj_lng, gcj_lat)
    # print("经纬度:", wgs84_lng, wgs84_lat)

    # 坐标系转换函数测试
    transfer('company_home.xlsx', 'gcj02_to_wgs84', 'GCJ02_route_points', 'WGS84_route_points')
