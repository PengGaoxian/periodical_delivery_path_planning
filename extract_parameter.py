import openpyxl
import numpy as np
# 获取编码的数字边界
def getEncodeNumBoundary(filePath, tableName):
    # 打开excel文件
    wb = openpyxl.load_workbook(filePath)
    # 获取sheet
    sheet = wb[tableName]
    # 遍历sheet中的第每一列
    encodeNumBoundary_list = []
    for index in range(1, sheet.max_column+1):
        # 遍历sheet中的第index列
        count = 0
        for cell in sheet.iter_rows(min_row=2, min_col=index, max_col=index):
            if cell[0].value != None:
                count += 1
            else:
                break
        encodeNumBoundary_list.append(count)
    return encodeNumBoundary_list

# 获取车辆的通勤时间（commuting_route.xlsx文件的commuting_route_grid表）
def getVehicleCommutingTime(filePath, tableName):
    # 打开filePath的tableName表
    wb = openpyxl.load_workbook(filePath)
    sheet = wb[tableName]
    # 存储车辆通勤时间
    vehicleCommutingTime_list = []
    # 遍历sheet中的第三列数据
    for cell in sheet.iter_rows(min_row=2, min_col=3, max_col=3):
        vehicleCommutingTime_list.append(cell[0].value)

    np_vehicleCommutingTime_list = np.array(vehicleCommutingTime_list)
    return np_vehicleCommutingTime_list / 60 # 表中数据单位为秒，转换为分钟

# 获取物流任务的信息（logistics_task_10.xlsx文件的WGS_logistics_task_10表）
def getLogisticsTaskInfo(filePath, tableName):
    # 打开filePath的tableName表
    wb = openpyxl.load_workbook(filePath)
    sheet = wb[tableName]
    # 存储物流任务的信息
    logisticsTaskInfo_list = []
    # 遍历sheet中的第八列到第十列数据（重量、起点、终点）
    for cell in sheet.iter_rows(min_row=2, min_col=8, max_col=10):
        logisticsTaskInfo_list.append([cell[0].value, cell[1].value, cell[2].value])
    return logisticsTaskInfo_list

# 获取物流任务配送路径中的接力点
def getTransPoints(filePath, tableName):
    # 打开filePath的tableName表
    wb = openpyxl.load_workbook(filePath)
    sheet = wb[tableName]
    # 存储物流任务配送路径中的接力点
    transPoints_list = []
    # 遍历sheet中的第一列数据
    for cell in sheet.iter_rows(min_row=2, min_col=2, max_col=2):
        transPoints_list.append(cell[0].value)
    return transPoints_list

# 计算车辆序列接力的最佳transPoint
def getRelayPoint(filePath, tableName, comp_3d_array):
    from A_star import calculate_heuristic
    transPoints_list = getTransPoints(filePath, tableName)
    relay4comp_2d_array = [] # 用于存储车辆序列组合接力点的列表
    # 遍历车辆序列组合comp_3d_array中的每一辆车
    for vehSeq4ST_2d_array in comp_3d_array:
        relay4ST_1d_array = [] # 用于存储车辆序列间接力点的列表
        if len(vehSeq4ST_2d_array) == 1:
            pass
        else:
            # 遍历车辆序列vehSeq4ST_2d_array中的相邻的两个两辆车
            for index in range(len(vehSeq4ST_2d_array)-1):
                # 获取相邻的两辆车
                vehPrev_1d_array = vehSeq4ST_2d_array[index][1:]
                vehNext_1d_array = vehSeq4ST_2d_array[index+1][1:]
                # 获取vehPrev_1d_array和vehNext_1d_array的交集
                vehPrevNext_1d_array = list(set(vehPrev_1d_array).intersection(set(vehNext_1d_array)))
                if vehPrevNext_1d_array == []:
                    vehPrevNext_1d_array = [vehPrev_1d_array[-1], vehNext_1d_array[0]]
                # 获取vehPrevNext_1d_array与transPoints_list的交集
                tmp_list = list(set(vehPrevNext_1d_array).intersection(set(transPoints_list)))
                # 如果tmp_list不为空，则返回tmp_list中的第一个元素
                if len(tmp_list) != 0:
                    relay4ST_1d_array.append(tmp_list[-1])
                # 如果tmp_list为空，则遍历transPoints_list找到与vehPrevNext_1d_array[0]和vehPrevNext_1d_array[-1]最近的点
                else:
                    shortest_distance1 = 100
                    shortest_grid1 = None
                    shortest_distance2 = 100
                    shortest_grid2 = None
                    # 计算transPoints_list中到vehPrevNext_1d_array[0]的最短距离
                    for transPoint in transPoints_list:
                        distance1 = calculate_heuristic(vehPrevNext_1d_array[0], transPoint)
                        distance2 = calculate_heuristic(vehPrevNext_1d_array[-1], transPoint)
                        if distance1 < shortest_distance1:
                            shortest_distance1 = distance1
                            shortest_grid1 = transPoint
                        if distance2 < shortest_distance2:
                            shortest_distance2 = distance2
                            shortest_grid2 = transPoint

                    # 比较shortest_distance1和shortest_distance2的大小
                    if shortest_distance1 < shortest_distance2:
                        relay4ST_1d_array.append(shortest_grid1)
                    else:
                        relay4ST_1d_array.append(shortest_grid2)
        relay4comp_2d_array.append(relay4ST_1d_array)
    return relay4comp_2d_array

            
# 获取候选车辆序列集
def getCandidateVehicleSeqSet(filepath, table):
    # 从candidate_commuting_route_set.xlsx文件的logistics_task表中获取候选车辆序列集
    import openpyxl
    # 打开文件
    wb = openpyxl.load_workbook(filepath)
    # 获取表
    sheet = wb[table]
    # 获取列数
    num_cols = sheet.max_column
    # 候选车辆序列集
    vehicleSeq4AllTask_4d_array= []
    # 遍历表，从第一列开始遍历列  
    for i in range(1, num_cols + 1):
        # 车辆序列
        vehicleSeq4SingleTask_3d_array = []
        # 从第二行开始遍历所有行，第1行为表头
        for cell in sheet[openpyxl.utils.get_column_letter(i)][1:]:
            # 如果cell.value不为空
            if cell.value != None:
                vehicleSeq4SingleTask_2d_array = eval(cell.value)
                if vehicleSeq4SingleTask_2d_array != []:
                    # 添加到车辆序列到集合中
                    vehicleSeq4SingleTask_3d_array.append(vehicleSeq4SingleTask_2d_array)
            else:
                break
        # 添加到候选车辆序列集中
        vehicleSeq4AllTask_4d_array.append(vehicleSeq4SingleTask_3d_array) 

    return vehicleSeq4AllTask_4d_array




if __name__ == '__main__':
    filePath = 'dataSet/candidate_commuting_route_set.xlsx'
    tableName = 'logistics_task_10'
    comp_3d_array = [
        [['C14', 1079, 1080, 1081, 1026], ['C55', 1081, 1026, 971, 916, 915, 860]],
        [['C95', 946, 947, 948, 949, 950, 951], ['C64', 947, 948, 949, 950, 951, 952, 1007, 1008, 1009, 1010], ['C81', 1007, 1008, 1009, 1010, 955, 956, 957, 958, 959, 960, 1015, 1016, 1017], ['C42', 1015, 1016, 1017, 1018, 1019, 1020, 1021, 1076, 1077, 1078, 1079]],
        [['C40', 1716, 1717, 1718, 1719, 1720], ['C94', 1720, 1665, 1610, 1555], ['C95', 1555, 1500, 1445, 1390, 1335, 1280, 1225, 1226, 1171, 1116, 1061, 1006, 951, 950, 949, 948, 947, 946]],
        [['C84', 1548, 1603, 1658, 1713, 1714, 1715, 1716]],
        [['C38', 370, 369, 368, 367, 366, 365, 310, 309, 308, 307, 306, 305, 304, 303, 302, 301, 300, 299, 354, 353, 352, 351, 350, 349, 348, 347, 346, 345], ['C100', 353, 352, 351, 350, 349, 348, 347, 346, 345, 344, 343, 398, 453, 452, 451, 506, 561, 616, 671, 726, 781, 836, 835, 834, 833, 888, 887, 942, 997, 1052, 1107], ['C34', 1107, 1162, 1217, 1272, 1327, 1382, 1383, 1438, 1493, 1548]],
        [['C35', 1023, 968, 913, 858, 803, 804, 805, 860, 861, 862, 863, 864, 865, 866, 867], ['C88', 867, 812, 757, 702, 701, 646, 591, 536, 535, 480, 425, 370]],
        [['C39', 860, 859, 804, 803], ['C33', 803, 858, 913, 968, 1023]],
        [['C26', 1330, 1329, 1274, 1219, 1164, 1109, 1110, 1111, 1166, 1167, 1168, 1169, 1170, 1171, 1226, 1227, 1228, 1229, 1230], ['C98', 1109, 1110, 1111, 1166, 1167, 1168, 1169, 1170, 1171, 1226, 1227, 1228, 1229, 1230, 1231, 1232, 1233, 1234, 1235, 1236, 1237, 1238, 1239, 1240, 1241, 1296, 1297, 1298, 1299, 1300, 1301], ['C55', 1301, 1246, 1191, 1136, 1081, 1026, 971, 916, 915, 860]],
        [['C53', 1142, 1143, 1144, 1199, 1254, 1309, 1364], ['C98', 1364, 1363, 1362, 1361, 1306, 1305, 1304, 1303, 1302, 1301, 1300, 1299, 1298, 1297, 1296, 1241, 1240, 1239, 1238, 1237, 1236, 1235, 1234, 1233, 1232, 1231, 1230, 1229, 1228, 1227, 1226, 1171, 1170, 1169, 1168, 1167, 1166, 1111, 1110, 1109], ['C26', 1230, 1229, 1228, 1227, 1226, 1171, 1170, 1169, 1168, 1167, 1166, 1111, 1110, 1109, 1164, 1219, 1274, 1329, 1330]],
        [['C57', 860, 861, 862, 863, 864, 865, 866, 867, 922, 977, 1032, 1033, 1088, 1089, 1144], ['C42', 1143, 1142]]
    ]
    # code_list = [1,2,3,4,5,6,7,8,9,10]

    # code = encode(filePath, tableName, comp_3d_array)
    # print(code)

    # decode_list = decode(filePath, tableName, code_list)
    # print(decode_list)

    # encodeboundary = getEncodeNumBoundary(filePath, tableName)
    # print(encodeboundary)

    trans_points = getTransPoints('dataSet/trans_points.xlsx', '100%')
    print(len(trans_points))

    relay_points = getRelayPoint(comp_3d_array, trans_points)
    print(relay_points)