import openpyxl
# 更新area_grid_seqtable.xlsx文件，根据area_grid_seq表中的route_ID和grid_uid，生成area_grid_seq_new表
def area_grid_seq_deduplicate():
    # 读取xlsx文件
    workbook = openpyxl.load_workbook('area_grid_seqtable.xlsx')
    # 获取xlsx文件中的sheet
    source_sheet = workbook['area_grid_seq']
    target_sheet = workbook['area_grid_seq_deduplicate']
    # 遍历source_sheet中的每一行，从第二行开始
    route_ID_prev = 0
    grid_uid_prev = 0
    for i in range(2, source_sheet.max_row + 1):
        # 读取第5列的route_ID和第3列的grid_uid
        route_ID = source_sheet.cell(row=i, column=5).value
        grid_uid = source_sheet.cell(row=i, column=3).value

        if route_ID != route_ID_prev:
            # 将route_ID和grid_uid写入target_sheet中，以追加的形式写入，从第二行开始写入
            target_sheet.append([route_ID, grid_uid])

        elif route_ID == route_ID_prev:
            if grid_uid != grid_uid_prev:
                # 将route_ID和grid_uid写入target_sheet中，以追加的形式写入，从第二行开始写入
                target_sheet.append([route_ID, grid_uid])
        route_ID_prev = route_ID
        grid_uid_prev = grid_uid

    # 保存工作簿到xlsx文件中
    workbook.save('area_grid_seqtable.xlsx')
    # 关闭工作簿
    workbook.close()

# 更新area_grid_seqtable.xlsx文件，根据area_grid_seq_new表中的route_ID和grid_uid，生成area_grid_neighbor_weight表
def area_grid_neighbor_weight():
    # 读取xlsx文件
    workbook = openpyxl.load_workbook('area_grid_seqtable.xlsx')
    # 获取xlsx文件中的sheet
    source_sheet = workbook['area_grid_seq_new']
    target_sheet = workbook['neighbor_weight']
    # 遍历source_sheet中的每一行，从第二行开始
    route_ID_prev = 0
    grid_uid_prev = 0
    for i in range(2, source_sheet.max_row + 1):
        # 读取第5列的route_ID和第3列的grid_uid
        route_ID = source_sheet.cell(row=i, column=1).value
        grid_uid = source_sheet.cell(row=i, column=2).value

        if route_ID == route_ID_prev:
            # 将route_ID和grid_uid写入target_sheet中，以追加的形式写入，从第二行开始写入
            target_sheet.append([route_ID, grid_uid_prev, grid_uid, 1])

        route_ID_prev = route_ID
        grid_uid_prev = grid_uid

    # 保存工作簿到xlsx文件中
    workbook.save('area_grid_seqtable.xlsx')
    # 关闭工作簿
    workbook.close()

# 查找area_grid_Neighbors.xlsx文件中area_grid_Neighbors表中第二列和第三列的值，如果第二列的值为2，第三列的值为3，则将第四列的值改为1
def update_neighbor_weight():
    # 读取xlsx文件
    workbook1 = openpyxl.load_workbook('area_grid_Neighbors.xlsx')
    workbook2 = openpyxl.load_workbook('area_grid_seqtable.xlsx')    
    # 获取xlsx文件中的sheet
    source_sheet1 = workbook1['area_grid_Neighbors']
    source_sheet2 = workbook2['neighbor_weight']
    target_sheet = workbook1['area_grid_Neighbors']
    # 遍历source_sheet中的每一行，从第二行开始
    for i in range(2, source_sheet1.max_row + 1):
        # 读取第二、三列的值
        value_2 = source_sheet1.cell(row=i, column=2).value
        value_3 = source_sheet1.cell(row=i, column=3).value
        # 遍历source_sheet中的每一行，从第二行开始
        for j in range(2, source_sheet2.max_row + 1):
            value_2_compare = source_sheet2.cell(row=j, column=2).value
            value_3_compare = source_sheet2.cell(row=j, column=3).value
            
            # 如果第二列的值为2，第三列的值为3，则将第四列的值改为1
            if (value_2 == value_2_compare and value_3 == value_3_compare) or (value_2 == value_3_compare and value_3 == value_2_compare):
                target_sheet.cell(row=i, column=7).value = source_sheet2.cell(row=j, column=4).value
        print(i)
    # 保存工作簿到xlsx文件中
    workbook1.save('area_grid_Neighbors.xlsx')
    # 关闭工作簿
    workbook1.close()
    workbook2.close()